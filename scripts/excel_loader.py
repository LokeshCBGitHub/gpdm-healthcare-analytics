from __future__ import annotations

import logging
import os
import re
import sqlite3
from typing import Iterable, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
    _HAS_OPENPYXL = True
except Exception:
    openpyxl = None
    InvalidFileException = Exception
    _HAS_OPENPYXL = False

_log = logging.getLogger("gpdm.excel_loader")


_VALID_IDENT = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _sanitize_id(name: str) -> Optional[str]:
    if not name:
        return None
    cleaned = re.sub(r"[^A-Za-z0-9_]+", "_", str(name)).strip("_")
    if not cleaned:
        return None
    if not _VALID_IDENT.match(cleaned):
        cleaned = "_" + cleaned
    return cleaned[:60]


def _is_blank_row(row: Iterable) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


def _coerce(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, bool):
        return "1" if v else "0"
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    return str(v)


def load_excel_to_sqlite(
    data_dir: str,
    conn: Optional[sqlite3.Connection] = None,
) -> sqlite3.Connection:
    if conn is None:
        conn = sqlite3.connect(":memory:")

    if not _HAS_OPENPYXL:
        _log.warning("openpyxl not installed; Excel files in %s will be skipped",
                      data_dir)
        return conn
    if not os.path.isdir(data_dir):
        return conn

    max_rows     = int(os.environ.get("GPDM_EXCEL_MAX_ROWS", "1000000"))
    skip_hidden  = os.environ.get("GPDM_EXCEL_SKIP_HIDDEN", "1") == "1"
    debug        = os.environ.get("GPDM_EXCEL_DEBUG") == "1"

    for fname in sorted(os.listdir(data_dir)):
        low = fname.lower()
        if not (low.endswith(".xlsx") or low.endswith(".xlsm")):
            continue
        fpath = os.path.join(data_dir, fname)
        base = os.path.splitext(fname)[0]
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except (InvalidFileException, Exception) as exc:
            _log.warning("Skipping %s: %s", fname, exc)
            continue
        try:
            usable_sheets = []
            for sheet in wb.worksheets:
                if skip_hidden and sheet.sheet_state != "visible":
                    continue
                if sheet.max_row is not None and sheet.max_row <= 0:
                    continue
                usable_sheets.append(sheet)
            multi = len(usable_sheets) > 1
            for sheet in usable_sheets:
                tname_raw = f"{base}__{sheet.title}" if multi else base
                table = _sanitize_id(tname_raw)
                if table is None:
                    continue
                try:
                    _ingest_sheet(conn, sheet, table, max_rows=max_rows)
                    if debug:
                        _log.info("Loaded %s::%s → %s", fname, sheet.title, table)
                except Exception as exc:
                    _log.warning("Failed to ingest %s::%s — %s",
                                  fname, sheet.title, exc)
        finally:
            wb.close()
    conn.commit()
    return conn


def _ingest_sheet(conn: sqlite3.Connection, sheet, table: str,
                   max_rows: int) -> None:
    rows_iter = sheet.iter_rows(values_only=True)

    headers: List[str] = []
    for raw in rows_iter:
        if _is_blank_row(raw):
            continue
        for i, h in enumerate(raw):
            hs = _coerce(h)
            if hs is None or not hs.strip():
                headers.append(f"col_{i+1}")
            else:
                headers.append(hs.strip())
        break
    if not headers:
        return

    seen = {}
    clean_headers: List[str] = []
    for h in headers:
        key = h
        c = seen.get(key, 0)
        if c > 0:
            clean_headers.append(f"{h}_{c+1}")
        else:
            clean_headers.append(h)
        seen[key] = c + 1

    col_defs = ", ".join(f'"{h}" TEXT' for h in clean_headers)
    placeholders = ", ".join(["?"] * len(clean_headers))

    conn.execute(f'DROP TABLE IF EXISTS "{table}"')
    conn.execute(f'CREATE TABLE "{table}" ({col_defs})')

    batch: List[Tuple] = []
    count = 0
    insert_sql = f'INSERT INTO "{table}" VALUES ({placeholders})'
    for raw in rows_iter:
        if count >= max_rows:
            break
        if _is_blank_row(raw):
            continue
        row = [(_coerce(v) if v is not None else None) for v in raw]
        if len(row) < len(clean_headers):
            row = row + [None] * (len(clean_headers) - len(row))
        elif len(row) > len(clean_headers):
            row = row[:len(clean_headers)]
        batch.append(tuple(row))
        count += 1
        if len(batch) >= 1000:
            conn.executemany(insert_sql, batch); batch.clear()
    if batch:
        conn.executemany(insert_sql, batch)


def load_any_to_sqlite(
    data_dir: str,
    conn: Optional[sqlite3.Connection] = None,
) -> sqlite3.Connection:
    try:
        from recommendation_engine import load_csv_to_sqlite as _csv
    except Exception:
        _csv = None
    if _csv is not None:
        conn = _csv(data_dir)
    return load_excel_to_sqlite(data_dir, conn=conn)


if __name__ == "__main__":
    import argparse, json
    logging.basicConfig(level=logging.INFO)
    ap = argparse.ArgumentParser()
    ap.add_argument("--dir", required=True)
    ap.add_argument("--db",  default=":memory:")
    args = ap.parse_args()
    conn = sqlite3.connect(args.db)
    load_excel_to_sqlite(args.dir, conn)
    tables = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type IN ('table','view')").fetchall()]
    out = {}
    for t in tables:
        n = conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
        cols = [r[1] for r in conn.execute(f'PRAGMA table_info("{t}")').fetchall()]
        out[t] = {"rows": n, "columns": cols}
    print(json.dumps(out, indent=2))
